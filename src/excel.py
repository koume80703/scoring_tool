import re
import os
import openpyxl as px
from openpyxl.styles import PatternFill

from directory import Directory


class Excel:
    def __init__(self, wb_path):
        self.wb_path = wb_path
        self.workbook = px.load_workbook(wb_path)

        self.worksheet = self.workbook.worksheets[0]

    def write_score(
        self, grade: int, cls: int, num: int, score: int, comment: str
    ) -> None:
        """指定のエクセルファイルに採点結果を書き込むメソッド

        Args:
            grade (int): 採点対象の学年
            cls (int): 採点対象のクラス
            num (int): 採点対象の番号
            score (int): 点数
            comment (str): コメント
        """

        for row in self.worksheet.iter_rows(min_row=2):
            grade_xls = int(row[1].value)
            cls_xls = int(row[2].value)
            num_xls = int(row[3].value)

            if (grade, cls, num) == (grade_xls, cls_xls, num_xls):
                row[7].value = str(score) + "点: " + comment

    def save_workbook(self) -> None:
        """ワークブックを保存するメソッド"""

        self.workbook.save(self.wb_path)

    def get_score(self) -> list[tuple[int, str]]:
        scores = []
        for row in self.worksheet.iter_rows(min_row=2):
            feedback_cmt = row[7].value
            if feedback_cmt is None:
                score = None
            else:
                score = feedback_cmt.split(":")[0]
                score = int(re.sub(r"\D", "", score))
            scores.append(
                (
                    row[0].value,
                    int(row[1].value),
                    int(row[2].value),
                    int(row[3].value),
                    row[4].value,
                    row[5].value,
                    row[6].value,
                    score,
                )
            )
        else:
            return scores


def main():
    root_path = os.path.dirname(os.getcwd())
    excel_file_path = os.path.join(root_path, "xls_scored/")
    xls_files = Directory.get_all_file(excel_file_path)

    all_scores = [None for _ in range(len(xls_files))]
    for f_path in xls_files:
        i = int(os.path.splitext(f_path)[0].split("_")[-1]) - 1
        all_scores[i] = Excel(os.path.join(excel_file_path, f_path)).get_score()

    scores_fe_students = []
    for info in all_scores[0]:
        scores_fe_students.append(list(info))
    for sheet in all_scores[1:]:
        for info in sheet:
            grade = info[1]
            cls = info[2]
            num = info[3]
            score = info[7]

            for sfs in scores_fe_students:
                if sfs[1] == grade and sfs[2] == cls and sfs[3] == num:
                    sfs.append(score)
                    break
            else:
                print("not found:", grade, cls, num)

    wb = px.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "所属"
    sheet.cell(row=1, column=2).value = "学年"
    sheet.cell(row=1, column=3).value = "組"
    sheet.cell(row=1, column=4).value = "番号"
    sheet.cell(row=1, column=5).value = "氏名"
    sheet.cell(row=1, column=6).value = "カナ氏名"
    sheet.cell(row=1, column=7).value = "学生番号"
    for i in range(1, 12):
        sheet.cell(row=1, column=i + 7).value = "提出課題" + str(i)

    for i in range(len(scores_fe_students)):
        sfs = scores_fe_students[i]
        for j in range(len(sfs)):
            if sfs[j] is None:
                sheet.cell(row=i + 2, column=j + 1).value = 0
                sheet.cell(row=i + 2, column=j + 1).fill = PatternFill(
                    fgColor="FFFF00", bgColor="FFFF00", fill_type="solid"
                )
            else:
                sheet.cell(row=i + 2, column=j + 1).value = sfs[j]

    wb.save("../xls_scored/採点結果まとめ.xlsx")


if __name__ == "__main__":
    main()
